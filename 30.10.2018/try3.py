import numpy as np
import math
import openpyxl
import matplotlib.pyplot as pl
 
class Kmeans(object):
def __init__(self, dataset, num_clusters, stop_criterion, max_iter):
self.dataset = dataset
self.num_clusters = num_clusters
self.stop_criterion = stop_criterion
self.iteration = 0
self.clusters = {i: [] for i in range(self.num_clusters)}
self.centroids = np.ndarray((num_clusters, dataset.shape[1]))
self.old_centroids = []
self.max_iter = max_iter
self.sum_intra_cluster_dist = []
self.center_mean = np.ndarray((num_clusters, dataset.shape[1]))
self.intra_cluster_dist = []
 
def assign_data_to_cluster(self):
self.clusters = {i: [] for i in range(self.num_clusters)}
for datapoint in self.dataset:
mean_index = \
min([(m[0], np.linalg.norm(datapoint - self.centroids[m[0]])) for m in enumerate(self.centroids)],
key=lambda t: t[1])[0]
try:
self.clusters[mean_index].append(datapoint)
except KeyError:
self.clusters[mean_index] = [datapoint]
 
for key, cluster in self.clusters.items():
if not cluster:
cluster.append(self.dataset[np.random.randint(0, len(self.dataset), size=1)].flatten().tolist())
 
def assign_data_to_cluster_modified(self):
# clusters dictionary must empty in order to repopulate
self.clusters = {i: [] for i in range(self.num_clusters)}
for datapoint in self.dataset:
mean_index = \
min([(m[0], np.linalg.norm(datapoint - self.center_mean[m[0]])) for m in enumerate(self.center_mean)],
key=lambda t: t[1])[0]
try:
self.clusters[mean_index].append(datapoint)
except KeyError:
self.clusters[mean_index] = [datapoint]
 
for key, cluster in self.clusters.items():
if not cluster:
cluster.append(self.dataset[np.random.randint(0, len(self.dataset), size=1)].flatten().tolist())
 
def calc_intra_cluster_dist(self):
total_sum = 0
local_sum = 0
self.intra_cluster_dist = []
for cluster, samples in self.clusters.items():
centroid = self.centroids[cluster]
local_sum = 0
for i in range(len(samples)):
total_sum += np.linalg.norm(samples[i] - centroid)
local_sum += np.linalg.norm(samples[i] - centroid)
self.intra_cluster_dist.append(total_sum)
return total_sum
 
def re_calc_centroids(self):
 
self.old_centroids.append(self.centroids)
centers = np.ndarray(shape=self.centroids.shape)
for key, datapoints in self.clusters.items():
temp_mean = []
temp_sam = np.array(datapoints)
 
for i in range(self.dataset.shape[1]):
temp_mean.append(sum(temp_sam[:, i]) / temp_sam.shape[0])
centers[key] = np.array(temp_mean)
self.center_mean[key] = centers[key]
 
for i in range(centers.shape[0]):
distances = [np.linalg.norm(centers[i] - sample) for sample in self.clusters[i]]
new_centroid = distances.index(min(distances))
self.centroids[i] = self.clusters[i][new_centroid]
 
def has_converged(self):
if self.iteration > self.max_iter:
return True
elifself.iteration <= 3:
return False
else:
return math.isclose(self.sum_intra_cluster_dist[-1], self.sum_intra_cluster_dist[-2],
abs_tol=self.stop_criterion)
 
def has_converged_2(self):
if self.iteration > self.max_iter:
return True
elifself.iteration <= 2:
return False
else:
return np.array_equal(self.old_centroids[self.iteration - 1], self.centroids)
 
def randomize_centroids(self):
self.centroids = self.dataset[np.random.choice(self.dataset.shape[0], size=self.num_clusters, replace=False), :]
 
def crete_clusters(self):
self.randomize_centroids()
while not self.has_converged():
self.assign_data_to_cluster()
self.re_calc_centroids()
self.sum_intra_cluster_dist.append(self.calc_intra_cluster_dist())
self.iteration += 1
 
class Bisect_Kmeans():
def __init__(self, dataset, num_clusters, stop_criterion, num_trials, max_iter):
self.dataset = dataset
self.num_clusters = num_clusters
self.stop_criterion = stop_criterion
self.cluster_cnt = 0
self.clusters = {}
self.centroids = np.ndarray((num_clusters, dataset.shape[1]))
self.old_centroids = []
self.max_iter = max_iter
self.sum_intra_cluster_dist = []
self.num_trials = num_trials
self.bisect_iter = 0
self.list_of_trial_clusters = []
self.list_of_trial_SSE = []
self.final_bisection_result = []
 
def perform_clustering(self):
kmeans_initial = Kmeans(self.dataset, 1, self.stop_criterion, self.max_iter)
kmeans_initial.crete_clusters()
self.clusters.update(kmeans_initial.clusters)
self.sum_intra_cluster_dist.append(kmeans_initial.intra_cluster_dist)
self.cluster_cnt += 1
while self.cluster_cnt < self.num_clusters:
if len(self.sum_intra_cluster_dist) > 1:
max_SEE_index = max(enumerate(self.sum_intra_cluster_dist), key=lambda k: k[1])[0]
else:
max_SEE_index = 0
new_data = np.array(self.clusters[max_SEE_index])
self.list_of_trial_SSE = []
self.list_of_trial_clusters = []
self.final_bisection_result = []
for i in range(self.num_trials):
kmeans = Kmeans(new_data, 2, self.stop_criterion, self.max_iter)
kmeans.crete_clusters()
for value in kmeans.clusters.values():
self.list_of_trial_clusters.append(value)
for j in range(kmeans.num_clusters):
self.list_of_trial_SSE.append(kmeans.intra_cluster_dist[j])
sum_pair_SSE = [self.list_of_trial_SSE[x] + self.list_of_trial_SSE[x + 1] for x in
range(int(len(self.list_of_trial_SSE) / 2))]
min_SSE_index = min(enumerate(sum_pair_SSE), key=lambda m: m[1])[0]
self.final_bisection_result.append(self.list_of_trial_clusters[2 * min_SSE_index])
self.final_bisection_result.append(self.list_of_trial_clusters[2 * min_SSE_index + 1])
self.clusters.pop(max_SEE_index)
self.sum_intra_cluster_dist.remove(self.sum_intra_cluster_dist[max_SEE_index])
self.sum_intra_cluster_dist.append(self.list_of_trial_SSE[2 * min_SSE_index])
self.sum_intra_cluster_dist.append(self.list_of_trial_SSE[2 * min_SSE_index + 1])
for _cluster in self.final_bisection_result:
for m in range(len(self.clusters) + 2):
if m not in self.clusters.keys():
self.clusters.update({m: _cluster})
break
self.cluster_cnt += 1
 
def read_in_data():
dataset = openpyxl.load_workbook('TwoDimensionalContinuousData.xlsx')
ws = dataset.get_active_sheet()
data = np.zeros([ws.max_row, ws.max_column])
for i in range(2, ws.max_row + 1):
for j in range(1, ws.max_column + 1):
data[i - 2, j - 1] = ws.cell(row=i, column=j).value
return data
 
def plot_data(original_data, clustered_data):
BEFORE = pl.figure(figsize=(5, 4))
pl.plot(original_data[:, 0], original_data[:, 1], 'o', label=r'Original Data')
# BEFORE.show()
pl.xlabel(r'X_1', fontsize=10)
pl.ylabel(r'X_2', fontsize=10)
pl.legend(loc='upper right', fontsize=10)
pl.xlim([-1, 6])
pl.ylim([-1, 6])
BEFORE.savefig("Original Dataset", dpi=200)
 
colors = ['red', 'green', 'indigo', 'black', 'royalblue', 'brown', 'orange', 'lime', 'darkkhaki', 'cyan', 'gold']
AFTER = pl.figure(figsize=(5, 4))
pl.xlim([-1, 6])
pl.ylim([-1, 6])
for key in clustered_data.clusters.keys():
x = [item[0] for item in clustered_data.clusters[key]]
y = [item[1] for item in clustered_data.clusters[key]]
pl.scatter(x, y, color=colors[key % len(colors)], marker='o', label=r'Cluster {}'.format(key))
pl.xlabel(r'X_1', fontsize=10)
pl.ylabel(r'X_2', fontsize=10)
pl.legend(loc='upper right', fontsize=10)
pl.show()
AFTER.savefig("Data after clustering{}".format(clustered_data.num_clusters), dpi=200)
 
if __name__ == '__main__':
data = read_in_data()
total_SSE = []
k = 11
for i in range(1, k):
bisect_kmeans = Bisect_Kmeans(dataset=data, num_clusters=i, stop_criterion=0, num_trials=10, max_iter=100000)
bisect_kmeans.perform_clustering()
plot_data(data, bisect_kmeans)
total_SSE.append(np.sum(bisect_kmeans.sum_intra_cluster_dist))
 
elbow_plot = pl.figure(figsize=(5, 5))
pl.legend(loc='upper right', fontsize=10)
pl.plot(range(1, k), total_SSE, '-k', label=r'Elbow method')
pl.xlabel(r'Number of clusters', fontsize=10)
pl.ylabel(r'Total SSE', fontsize=10)
pl.show()
elbow_plot.savefig("Elbow method chart", dpi=200)